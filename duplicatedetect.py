import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog
from thefuzz import fuzz  # Fuzzy matching library


def highlight_similar_names(file_path, similarity_threshold=80):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Select the active sheet

    # Identify the "ItemName" column
    header = [cell.value for cell in ws[1]]  # Read first row as headers
    if "ItemName" not in header:
        print("Error: 'ItemName' column not found.")
        return

    itemname_col = header.index("ItemName") + 1  # Convert to 1-based index

    # Read all item names with their row numbers
    item_names = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=itemname_col).value
        if name:
            item_names[row] = (
                name.strip().lower()
            )  # Store row number and lowercase name

    # Compare all item names to find similar ones
    rows_to_highlight = set()
    item_keys = list(item_names.keys())

    for i in range(len(item_keys)):
        for j in range(i + 1, len(item_keys)):  # Avoid duplicate comparisons
            name1 = item_names[item_keys[i]]
            name2 = item_names[item_keys[j]]
            similarity_score = fuzz.ratio(name1, name2)  # Compute similarity

            if similarity_score >= similarity_threshold:
                rows_to_highlight.add(item_keys[i])
                rows_to_highlight.add(item_keys[j])

    # Highlight similar names in yellow
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    for row in rows_to_highlight:
        ws.cell(row=row, column=itemname_col).fill = (
            yellow_fill  # Highlight ItemName column
        )

    # Save the modified workbook
    output_file = file_path.replace(".xlsx", "_highlighted.xlsx")
    wb.save(output_file)
    print(f"✅ Similar item names highlighted in 'ItemName'.\nSaved as: {output_file}")


def upload_file():
    # Open a file dialog for the user to choose an Excel file
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        print(f"📂 File selected: {file_path}")
        highlight_similar_names(file_path)


# Run the file upload function
upload_file()
